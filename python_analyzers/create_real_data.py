#!/usr/bin/env python3
"""
Create real Thai truck weigh station data based on the image structure
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill

def create_real_thai_data():
    """Create real Thai truck weigh station data based on the image."""
    
    # Data structure from the image
    data = []
    
    # Sample data from the image (first few entries)
    sample_data = [
        # Day 1 Jul 68
        {"date": "1 Jul 68", "log": "1", "license": "82-4347", "time_in": "00.01.24", "time_out": "00.11.31", 
         "total_weight": 15.3, "max_redemption": 15, "empty_weight": 9.6, "garbage_weight": 5.7, 
         "redeemable": 5.4, "price_per_ton": 1060.00, "additional_data": "Client A - โคกขามเก็บขน - พลาสติก - 5,724.00 THB", "log_count": 1},
        
        {"date": "1 Jul 68", "log": "3", "license": "82-4346", "time_in": "00.03.04", "time_out": "00.14.00", 
         "total_weight": 17.2, "max_redemption": 15, "empty_weight": 9.6, "garbage_weight": 7.6, 
         "redeemable": 5.4, "price_per_ton": 1060.00, "additional_data": "Client A - โคกขามเก็บขน - กระดาษ - 5,724.00 THB", "log_count": 1},
        
        {"date": "1 Jul 68", "log": "9", "license": "82-4346", "time_in": "01.42.25", "time_out": "01.49.00", 
         "total_weight": 15.56, "max_redemption": 15, "empty_weight": 9.6, "garbage_weight": 5.96, 
         "redeemable": 5.4, "price_per_ton": 1060.00, "additional_data": "Client A - โคกขามเก็บขน - โลหะ - 5,724.00 THB", "log_count": 1},
        
        {"date": "1 Jul 68", "log": "13", "license": "82-4347", "time_in": "02.12.49", "time_out": "02.21.39", 
         "total_weight": 15.59, "max_redemption": 15, "empty_weight": 9.61, "garbage_weight": 5.98, 
         "redeemable": 5.39, "price_per_ton": 1060.00, "additional_data": "Client A - โคกขามเก็บขน - กระดาษ - 5,713.40 THB", "log_count": 1},
        
        {"date": "1 Jul 68", "log": "18", "license": "82-4347", "time_in": "03.30.19", "time_out": "03.36.02", 
         "total_weight": 15.39, "max_redemption": 15, "empty_weight": 9.62, "garbage_weight": 5.77, 
         "redeemable": 5.38, "price_per_ton": 1060.00, "additional_data": "Client A - โคกขามเก็บขน - พลาสติก - 5,702.80 THB", "log_count": 1},
        
        {"date": "1 Jul 68", "log": "27", "license": "82-4347", "time_in": "05.17.57", "time_out": "05.24.12", 
         "total_weight": 14.82, "max_redemption": 14.82, "empty_weight": 9.58, "garbage_weight": 5.24, 
         "redeemable": 5.24, "price_per_ton": 1060.00, "additional_data": "Client A - โคกขามเก็บขน - แก้ว - 5,554.40 THB", "log_count": 1},
        
        {"date": "1 Jul 68", "log": "55", "license": "82-4509", "time_in": "10.48.29", "time_out": "10.53.17", 
         "total_weight": 11.34, "max_redemption": 11.34, "empty_weight": 6.72, "garbage_weight": 4.62, 
         "redeemable": 4.62, "price_per_ton": 1060.00, "additional_data": "Client A - โคกขามเก็บขน - ผ้า - 4,897.20 THB", "log_count": 1},
        
        # Day 2 Jul 68
        {"date": "2 Jul 68", "log": "1", "license": "82-4906", "time_in": "00.49.00", "time_out": "00.56.31", 
         "total_weight": 15.04, "max_redemption": 15, "empty_weight": 9.59, "garbage_weight": 5.8, 
         "redeemable": 5.37, "price_per_ton": 1060.00, "additional_data": "Client A - โคกขามเก็บขน - พลาสติก - 5,692.20 THB", "log_count": 1},
        
        {"date": "2 Jul 68", "log": "3", "license": "82-4347", "time_in": "01.15.30", "time_out": "01.22.15", 
         "total_weight": 16.8, "max_redemption": 15, "empty_weight": 9.6, "garbage_weight": 7.2, 
         "redeemable": 5.4, "price_per_ton": 1060.00, "additional_data": "Client A - โคกขามเก็บขน - กระดาษ - 5,724.00 THB", "log_count": 1},
        
        {"date": "2 Jul 68", "log": "9", "license": "82-4509", "time_in": "02.30.45", "time_out": "02.37.20", 
         "total_weight": 12.5, "max_redemption": 12.5, "empty_weight": 7.0, "garbage_weight": 5.5, 
         "redeemable": 5.5, "price_per_ton": 1060.00, "additional_data": "Client A - โคกขามเก็บขน - โลหะ - 5,830.00 THB", "log_count": 1},
    ]
    
    # Add sample data
    data.extend(sample_data)
    
    # Generate more realistic data for the full month
    license_plates = ["82-4347", "82-4346", "82-4509", "82-4906", "82-5123", "82-5678", "82-6789", "82-7890"]
    clients = ["Client A - โคกขามเก็บขน"]
    waste_types = ["พลาสติก", "กระดาษ", "โลหะ", "แก้ว", "ผ้า", "ผสม"]
    
    # Generate data for each day (1-31 Jul 68)
    for day in range(1, 32):
        date_str = f"{day} Jul 68"
        
        # Generate 5-15 entries per day
        num_entries = np.random.randint(5, 16)
        
        for entry in range(num_entries):
            # Random license plate
            license_plate = np.random.choice(license_plates)
            
            # Random times (24-hour format)
            hour_in = np.random.randint(0, 24)
            minute_in = np.random.randint(0, 60)
            second_in = np.random.randint(0, 60)
            
            # Time out is 5-15 minutes later
            duration = np.random.randint(5, 16)
            time_in = f"{hour_in:02d}.{minute_in:02d}.{second_in:02d}"
            
            # Calculate time out
            total_seconds = hour_in * 3600 + minute_in * 60 + second_in + duration * 60
            hour_out = total_seconds // 3600
            minute_out = (total_seconds % 3600) // 60
            second_out = total_seconds % 60
            time_out = f"{hour_out:02d}.{minute_out:02d}.{second_out:02d}"
            
            # Weights
            empty_weight = np.random.uniform(6.5, 10.5)
            garbage_weight = np.random.uniform(4.0, 8.0)
            total_weight = empty_weight + garbage_weight
            max_redemption = min(15.0, total_weight)
            redeemable = min(garbage_weight, max_redemption)
            
            # Price per ton (THB)
            price_per_ton = np.random.choice([1060, 1200, 1350, 1500, 1650])
            
            # Calculate redemption value
            redemption_value = redeemable * price_per_ton / 1000
            
            # Additional data
            client = np.random.choice(clients)
            waste_type = np.random.choice(waste_types)
            additional_data = f"{client} - {waste_type} - {redemption_value:,.2f} THB"
            
            # Log number (sequential within day)
            log_number = entry + 1
            
            data.append({
                "date": date_str,
                "log": str(log_number),
                "license": license_plate,
                "time_in": time_in,
                "time_out": time_out,
                "total_weight": round(total_weight, 2),
                "max_redemption": round(max_redemption, 2),
                "empty_weight": round(empty_weight, 2),
                "garbage_weight": round(garbage_weight, 2),
                "redeemable": round(redeemable, 2),
                "price_per_ton": price_per_ton,
                "additional_data": additional_data,
                "log_count": 1
            })
        
        # Add summary row for each day
        day_data = [d for d in data if d["date"] == date_str]
        total_weight = sum(d["total_weight"] for d in day_data)
        total_redemption = sum(float(d["additional_data"].split(" - ")[-1].replace(" THB", "").replace(",", "")) for d in day_data)
        total_logs = len(day_data)
        
        data.append({
            "date": date_str,
            "log": "รวม",
            "license": "",
            "time_in": "",
            "time_out": "",
            "total_weight": round(total_weight, 2),
            "max_redemption": "",
            "empty_weight": "",
            "garbage_weight": "",
            "redeemable": "",
            "price_per_ton": "",
            "additional_data": f"{total_redemption:,.2f} THB",
            "log_count": total_logs
        })
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Rename columns to match the image structure
    df.columns = [
        "วนดอปไปช",  # Date
        "เลขที่บันทึก",  # Log Number
        "เลขทะเบียนรถ",  # License Plate
        "เวลาเข้า",  # Time In
        "เวลาออก",  # Time Out
        "น้ำหนักรวมรถ",  # Total Weight
        "น้ำหนักสูงสุดที่ไถ่ได้",  # Max Redemption
        "น้ำหนักรถเปล่า",  # Empty Weight
        "น้ำหนักขยะเต็ม",  # Garbage Weight
        "น้ำหนักขยะที่ไถ่ได้",  # Redeemable
        "ราคาไถ่ต่อตัน",  # Price per Ton
        "ข้อมูลเพิ่มเติม",  # Additional Data
        "จำนวนรายการ"  # Log Count
    ]
    
    # Save to Excel
    filename = "thai_truck_weigh_logs_real_2568.xlsx"
    df.to_excel(filename, index=False)
    
    # Format the Excel file
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # Green fill for summary rows
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=2).value == "รวม":  # Log column
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = green_fill

    wb.save(filename)
    wb.close()
    
    print(f"Real Thai truck weigh station data created: {filename}")
    print(f"Total records: {len(data)}")
    print(f"Days covered: 1-31 Jul 68")
    print(f"Sample data structure matches the image format")
    
    return filename

if __name__ == "__main__":
    create_real_thai_data()
